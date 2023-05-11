import funcoes as decoder
import time as timer
import serial as serial
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Caminhos.xlsx')
ws = wb.active

Serial = serial.Serial()
Serial.baudrate = 115200
Serial.port = 'COM3'
Serial.setRTS(False)
Serial.setDTR(False)
Serial.open()
print(Serial)
mensagem = ""
Serial.reset_input_buffer()
Serial.reset_output_buffer()
while Serial.isOpen():
    if Serial.in_waiting:
        recebido = Serial.readline()
        print(recebido)
    fala = input("O que você quer fazer?(Enviar/Sair): ")
    if fala ==  "Sair":
        Serial.write(b"[")
        Serial.close() 
    elif fala == "Enviar":
        resposta = b""
        Serial.write(b"r")
        mensagem = input("Digite sua mensagem: ")
        #Salvar a mensagem no excel
        ws.append(mensagem.split())
        mensagem = bytes(mensagem,'utf-8')
        mensagem = mensagem + b"\n"
        Serial.write(mensagem)
        Serial.flush()
        espera = True
        print("Espere...")
        timer.sleep(1)
        while Serial.inWaiting():
            resposta = resposta + Serial.readline(1)
        resposta = resposta.decode('ascii')
        print("Recebido: ")
        print(resposta)
        erro = input("Injetar erro?(s/n): ")
        if erro == 's':
            caminhoCorreto = decoder.Try(resposta)
            for linha in range(0,len(caminhoCorreto)):
                ws.append(caminhoCorreto[linha].tolist())
            resposta = decoder.Erro(resposta,50)
        print("Criando caminho...")
        bits = decoder.CriarBits(resposta)
        caminho = decoder.Viterbi(bits)
        print(caminho)
        print(len(caminho[0]),"bits recebidos")
        s = decoder.Reverter(caminho,len(caminho[0]))
        print(s)
        ws.append(s.split())
        for linha in range(0,len(caminho)):
            ws.append(caminho[linha].tolist())

wb.save('Caminhos.xlsx')
            